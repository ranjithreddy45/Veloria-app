#!/usr/bin/env python3
"""Normalize 'Vendor Rate Card.xlsx' -> vendor-import.json for the DB importer.
Deterministic. v2: parses the inclusions/menu text into structured Sections + Items
(FIXED / SINGLE_CHOICE / MULTI_CHOICE) and applies category overrides.
Run: python3 scripts/normalize-vendor-rate-card.py "<xlsx>" <out.json>
"""
import json, re, sys
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "/Users/ranjithreddy/Downloads/Vendor Rate Card.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/Users/ranjithreddy/Downloads/VeloriaApp/scripts/vendor-import.json"

# Sheet "Vendor Category" (col B) -> catalog category key.
CAT_KEY = {
    "Birthday Decor": "decor", "Decoration": "decor",
    "Pure Veg (North Indian)": "catering", "Non veg": "catering",
    "Photographer": "photography", "Photography": "photography",
    "Birthday Cake/Engagement Cake": "cakes",
    "Activities": "activities", "Live stall": "live_stalls",
}
# Per-package category overrides (the sheet lumps these under "Birthday Decor" but
# they are emcee/entertainment services). Keyed by (vendor, package name lower).
CAT_OVERRIDE = {
    ("Event Echoes(Bharathshankar)", "emcee games"): "emcee",
    ("Event Echoes(Bharathshankar)", "magic show"): "entertainment",
    ("Event Echoes(Bharathshankar)", "caricature duration"): "entertainment",
}
# Catalog categories that must exist (created if missing). decor/catering/photography/
# emcee/entertainment already ship in VENDOR_CATEGORIES; only these are new.
NEW_CATEGORIES = [
    {"key": "cakes", "label": "Cakes"},
    {"key": "activities", "label": "Activities"},
    {"key": "live_stalls", "label": "Live Stalls"},
]
KEY_TO_ENUM = {
    "decor": "DECORATION", "catering": "CATERING", "photography": "PHOTOGRAPHY",
    "cakes": "CATERING", "activities": "ENTERTAINMENT", "live_stalls": "ENTERTAINMENT",
    "emcee": "ENTERTAINMENT", "entertainment": "ENTERTAINMENT",
}
def price_unit(u):
    s = (u or "").strip().lower()
    return {"per event": "PER_EVENT", "per plate": "PER_PLATE", "each": "PER_PIECE",
            "per kg": "PER_PIECE", "per piece": "PER_PIECE"}.get(s, "PER_EVENT")

GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[0-9A-Z]{3}$")

def phone_str(v):
    if v is None: return None
    if isinstance(v, float): return str(int(v))
    return str(v).strip()

def num(v):
    return float(v) if isinstance(v, (int, float)) else None

# ---------- structured menu / inclusions parser ----------
def _clean(x):
    return re.sub(r"\s+", " ", x.strip().lstrip("*").strip())

def _split_choice(head):
    """'Welcome Drink (Any 1)' -> ('Welcome Drink', 1); '(3 - 5 options)' -> 3; else (head, None)."""
    m = re.search(r"\(\s*(?:any\s*)?(\d+)\s*(?:-\s*\d+)?\s*(?:options?|type|types)?\s*\)", head, re.I)
    if m:
        return re.sub(r"\s*\([^)]*\)\s*$", "", head).strip(), int(m.group(1))
    return head, None

def _fixed_section(title, items):
    return {"title": title, "items": [{"name": _clean(i), "type": "FIXED", "options": [], "chooseCount": None}
                                      for i in items if _clean(i)]}

def parse_structure(desc):
    """Return (sections, plain_description)."""
    if not desc:
        return [], None
    lines = [l.strip() for l in desc.replace("\r", "").split("\n")]
    lines = [l for l in lines if l]
    has_gt = any(l.startswith(">") for l in lines)
    has_star = any(l.startswith("*") for l in lines)

    if has_gt:
        raw_secs, cur = [], None
        for l in lines:
            if l.startswith(">"):
                title, choose = _split_choice(l[1:].strip())
                cur = {"title": title, "choose": choose, "lines": []}
                raw_secs.append(cur)
            else:
                body = l[1:].strip() if l.startswith("*") else l
                if cur is None:
                    cur = {"title": "Inclusions", "choose": None, "lines": []}
                    raw_secs.append(cur)
                cur["lines"].append(body)
        sections = []
        for s in raw_secs:
            if s["choose"] is not None:  # a choice course: gather options across its lines
                opts, seen = [], set()
                for ln in s["lines"]:
                    for o in re.split(r"\s*/\s*", ln):
                        o = _clean(o)
                        if o and o.lower() not in seen:
                            seen.add(o.lower()); opts.append(o)
                itype = "SINGLE_CHOICE" if s["choose"] == 1 else "MULTI_CHOICE"
                cc = None if s["choose"] == 1 else min(s["choose"], len(opts) or s["choose"])
                sections.append({"title": s["title"],
                                 "items": [{"name": s["title"], "type": itype, "options": opts, "chooseCount": cc}]})
            else:
                sections.append(_fixed_section(s["title"], s["lines"]))
        return [s for s in sections if s["items"]], None

    if has_star:
        return [_fixed_section("Inclusions", [l for l in lines if l.startswith("*")])], None

    # comma / plus list, or a plain phrase. Only treat it as an inclusions LIST when
    # it clearly is one (an "Inclusions:" prefix, a "+"-joined list, or >=3 parts);
    # otherwise a short 2-part phrase stays a plain description.
    joined = " ".join(lines)
    m = re.match(r"(?i)(inclusions?)\s*:?\s*(.*)", joined)
    is_incl = bool(m)
    body = m.group(2) if m else joined
    parts = [p.strip() for p in re.split(r"[+,]", body) if p.strip()]
    if len(parts) >= 3 or (len(parts) >= 2 and (is_incl or "+" in body)):
        return [_fixed_section("Inclusions", parts)], None
    return [], _clean(joined)

# ---------- workbook ----------
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["4. Rate Card"]
vendors, order = {}, []

def get_vendor(name, addr, email, phone, gstin):
    if name not in vendors:
        vendors[name] = {"name": name, "email": email, "phone": phone, "address": addr,
                         "city": "Bengaluru", "gstin": gstin, "categoryKeys": [], "packages": []}
        order.append(name)
    v = vendors[name]
    for k, val in (("email", email), ("phone", phone), ("address", addr), ("gstin", gstin)):
        if not v[k] and val: v[k] = val
    return v

for r in range(8, ws.max_row + 1):
    cat_raw, name = ws.cell(r, 2).value, ws.cell(r, 3).value
    if not name or not cat_raw:
        continue
    name = str(name).strip()
    addr = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else None
    email = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else None
    phone = phone_str(ws.cell(r, 6).value)
    pkg_name = str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else None
    unit_enum = price_unit(ws.cell(r, 8).value)
    vendor_rate = num(ws.cell(r, 9).value)
    customer_raw = ws.cell(r, 10).value
    desc = str(ws.cell(r, 13).value).strip() if ws.cell(r, 13).value else None
    l_val = ws.cell(r, 12).value
    gstin = str(l_val).strip() if (l_val and GSTIN_RE.match(str(l_val).strip())) else None

    key = CAT_OVERRIDE.get((name, (pkg_name or "").lower())) or CAT_KEY.get(cat_raw.strip())
    if not key:
        raise SystemExit(f"Row {r}: unmapped category {cat_raw!r}")

    v = get_vendor(name, addr, email, phone, gstin)
    if key not in v["categoryKeys"]:
        v["categoryKeys"].append(key)

    sections, plain = parse_structure(desc)

    def add_pkg(pn, cust, vrate, extra=None):
        secs, pd = sections, plain
        if extra:
            pd = (pd + " " if pd else "") + extra
        v["packages"].append({"name": pn, "category": key, "priceUnit": unit_enum,
                              "vendorPrice": vrate, "customerPrice": cust,
                              "description": pd, "sections": secs})

    if isinstance(customer_raw, str) and "/" in customer_raw and re.search(r"video", pkg_name or "", re.I):
        nums = [float(re.sub(r"[^0-9.]", "", p)) for p in customer_raw.split("/") if re.sub(r"[^0-9.]", "", p)]
        add_pkg("Photography", nums[0], vendor_rate)
        add_pkg("Photography + Videography", nums[1], vendor_rate)
    else:
        add_pkg(pkg_name, num(customer_raw), vendor_rate)

# stable category order for a nicer legacy-enum primary pick
PRIMARY_PREF = ["catering", "decor", "photography", "cakes", "emcee", "entertainment", "activities", "live_stalls"]
def primary_enum(keys):
    for k in PRIMARY_PREF:
        if k in keys:
            return KEY_TO_ENUM[k]
    return KEY_TO_ENUM[keys[0]]

out = {"categories": NEW_CATEGORIES, "vendors": []}
for n in order:
    v = vendors[n]
    out["vendors"].append({k: v[k] for k in ("name", "email", "phone", "address", "city", "gstin", "categoryKeys")}
                          | {"primaryEnum": primary_enum(v["categoryKeys"]), "packages": v["packages"]})

with open(OUT, "w") as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

npk = sum(len(v["packages"]) for v in out["vendors"])
nsec = sum(len(p["sections"]) for v in out["vendors"] for p in v["packages"])
print(f"Wrote {OUT}: {len(out['vendors'])} vendors, {npk} packages, {nsec} sections, {len(NEW_CATEGORIES)} new categories")
for v in out["vendors"]:
    struct = sum(1 for p in v["packages"] if p["sections"])
    print(f"  - {v['name']}  [{v['primaryEnum']}] cats={v['categoryKeys']} pkgs={len(v['packages'])} structured={struct}")
